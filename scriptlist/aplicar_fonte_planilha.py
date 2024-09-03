from openpyxl import load_workbook
from openpyxl.styles import Font

def aplicar_fonte_planilha(caminho_arquivo, nome_planilha, nome_fonte):
    try:
        wb = load_workbook(caminho_arquivo)
        sheet = wb[nome_planilha]

        fonte = Font(name=nome_fonte)

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = fonte

        wb.save(caminho_arquivo)
        print(f'Fonte "{nome_fonte}" aplicada à planilha "{nome_planilha}" em "{caminho_arquivo}".')

    except Exception as e:
        print(f'Erro ao processar o arquivo {caminho_arquivo}: {e}')

def main():
    caminho_arquivo = input("Digite o caminho do arquivo Excel: ").strip()
    nome_planilha = input("Digite o nome da planilha que deseja modificar: ").strip()
    nome_fonte = input("Digite o nome da fonte que deseja aplicar: ").strip()
    aplicar_fonte_planilha(caminho_arquivo, nome_planilha, nome_fonte)

if __name__ == "__main__":
    main()
