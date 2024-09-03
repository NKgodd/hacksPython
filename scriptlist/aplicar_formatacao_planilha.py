from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

def aplicar_formatacao_planilha(caminho_arquivo, nome_planilha, nome_fonte, tamanho_fonte, negrito, italico, sublinhado, alinhamento_horizontal, alinhamento_vertical, espaco_linhas):
    try:
        wb = load_workbook(caminho_arquivo)
        sheet = wb[nome_planilha]

        # Definir a fonte
        fonte = Font(name=nome_fonte, size=tamanho_fonte, bold=negrito, italic=italico, underline='single' if sublinhado else None)

        # Definir alinhamento
        alinhamento = Alignment(horizontal=alinhamento_horizontal, vertical=alinhamento_vertical)

        # Definir borda
        borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Aplicar formatação em cada célula
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = fonte
                cell.alignment = alinhamento
                cell.border = borda_fina

        # Ajustar altura das linhas para espaçamento
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = espaco_linhas

        wb.save(caminho_arquivo)
        print(f'Formatação aplicada à planilha "{nome_planilha}" em "{caminho_arquivo}".')

    except Exception as e:
        print(f'Erro ao processar o arquivo {caminho_arquivo}: {e}')

def main():
    caminho_arquivo = input("Digite o caminho do arquivo Excel: ").strip()
    nome_planilha = input("Digite o nome da planilha que deseja modificar: ").strip()
    nome_fonte = input("Digite o nome da fonte que deseja aplicar: ").strip()
    tamanho_fonte = int(input("Digite o tamanho da fonte: "))
    negrito = input("Deseja aplicar negrito? (s/n): ").strip().lower() == 's'
    italico = input("Deseja aplicar itálico? (s/n): ").strip().lower() == 's'
    sublinhado = input("Deseja aplicar sublinhado? (s/n): ").strip().lower() == 's'
    alinhamento_horizontal = input("Digite o alinhamento horizontal (center, left, right): ").strip().lower()
    alinhamento_vertical = input("Digite o alinhamento vertical (center, top, bottom): ").strip().lower()
    espaco_linhas = float(input("Digite a altura das linhas para espaçamento: "))

    aplicar_formatacao_planilha(caminho_arquivo, nome_planilha, nome_fonte, tamanho_fonte, negrito, italico, sublinhado, alinhamento_horizontal, alinhamento_vertical, espaco_linhas)

if __name__ == "__main__":
    main()
