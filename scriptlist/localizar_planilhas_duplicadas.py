import os
from openpyxl import load_workbook

def localizar_planilhas_duplicadas(pasta):
    planilhas_conteudo = {}
    planilhas_duplicadas = []

    for root, dirs, files in os.walk(pasta):
        for file in files:
            if file.endswith('.xlsx'):
                caminho_arquivo = os.path.join(root, file)
                try:
                    wb = load_workbook(caminho_arquivo, data_only=True)
                    
                    for sheet in wb.worksheets:
                        conteudo = []
                        for row in sheet.iter_rows():
                            conteudo.append(tuple(cell.value for cell in row))

                        conteudo = tuple(conteudo)
                        
                        if conteudo in planilhas_conteudo:
                            planilhas_duplicadas.append((caminho_arquivo, sheet.title))
                            planilhas_duplicadas.append(planilhas_conteudo[conteudo])
                        else:
                            planilhas_conteudo[conteudo] = (caminho_arquivo, sheet.title)
                    
                except Exception as e:
                    print(f'Erro ao abrir {caminho_arquivo}: {e}')

    if planilhas_duplicadas:
        print('Planilhas duplicadas encontradas:')
        for arquivo, planilha in planilhas_duplicadas:
            print(f'Arquivo: {arquivo}, Planilha: {planilha}')
    else:
        print('Nenhuma planilha duplicada encontrada.')

def main():
    caminho_pasta = input("Digite o caminho da pasta onde estão os arquivos Excel: ").strip()
    localizar_planilhas_duplicadas(caminho_pasta)

if __name__ == "__main__":
    main()
