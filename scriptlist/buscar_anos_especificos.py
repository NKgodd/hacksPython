import os
from openpyxl import load_workbook
import re

def listar_anos_em_excel(pasta, anos_correto):
    anos_correto = {str(ano) for ano in anos_correto}

    for root, dirs, files in os.walk(pasta):
        for file in files:
            if file.endswith('.xlsx'):
                caminho_arquivo = os.path.join(root, file)
                try:
                    wb = load_workbook(caminho_arquivo, data_only=True)
                    anos_encontrados = set()
                    posicoes_erradas = []

                    for sheet in wb.worksheets:
                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            for col_idx, cell in enumerate(row, start=1):
                                if isinstance(cell, str):
                                    encontrados = re.findall(r'\b(200\d|201\d|2020|2021)\b', cell)
                                    for ano in encontrados:
                                        anos_encontrados.add(ano)
                                        if ano not in anos_correto:
                                            posicoes_erradas.append(f'Sheet: {sheet.title}, Row: {row_idx}, Column: {col_idx} - Ano encontrado: {ano}')
                                elif isinstance(cell, int):
                                    ano = str(cell)
                                    anos_encontrados.add(ano)
                                    if ano not in anos_correto:
                                        posicoes_erradas.append(f'Sheet: {sheet.title}, Row: {row_idx}, Column: {col_idx} - Ano encontrado: {ano}')
                    
                    if posicoes_erradas:
                        print(f'\nArquivo: {caminho_arquivo}')
                        for posicao in posicoes_erradas:
                            print(f'  - {posicao}')
                    
                except Exception as e:
                    print(f'Erro ao abrir {caminho_arquivo}: {e}')

def main():
    caminho_pasta = input("Digite o caminho da pasta onde estão os arquivos Excel: ").strip()
    anos_input = input("Digite os anos corretos, separados por vírgula (por exemplo, 2000,2001,2002): ").strip()
    
    try:
        anos_correto = [int(ano.strip()) for ano in anos_input.split(',')]
    except ValueError:
        print("Por favor, insira uma lista válida de anos separados por vírgula.")
        return
    
    listar_anos_em_excel(caminho_pasta, anos_correto)

if __name__ == "__main__":
    main()
