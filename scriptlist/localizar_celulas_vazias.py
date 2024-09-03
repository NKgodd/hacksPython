import os
from openpyxl import load_workbook

def localizar_planilhas_vazias(pasta):
    for root, dirs, files in os.walk(pasta):
        for file in files:
            if file.endswith('.xlsx'):
                caminho_arquivo = os.path.join(root, file)
                try:
                    wb = load_workbook(caminho_arquivo, data_only=True)
                    planilhas_vazias = []

                    for sheet in wb.worksheets:
                        if all(cell.value is None or cell.value == '' for row in sheet.iter_rows() for cell in row):
                            planilhas_vazias.append(sheet.title)

                    if planilhas_vazias:
                        print(f'\nArquivo: {caminho_arquivo}')
                        print('Planilhas vazias encontradas:')
                        for planilha in planilhas_vazias:
                            print(f'  - {planilha}')
                    
                except Exception as e:
                    print(f'Erro ao abrir {caminho_arquivo}: {e}')

def main():
    caminho_pasta = r"C:\Users\Adm Manhã\Documents\arquivosxl"
    localizar_planilhas_vazias(caminho_pasta)

if __name__ == "__main__":
    main()
