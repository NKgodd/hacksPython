import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

def identificar_e_editar_titulos(sheet, novo_titulo):
    titulos = []
    for row in sheet.iter_rows(min_row=1, max_row=1):  # Supondo que o título esteja na primeira linha
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.isupper():
                titulos.append((cell.coordinate, cell.value))
                cell.value = novo_titulo  # Edita o título com o novo valor
    return titulos

def verificar_datas(sheet):
    datas = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "data" in cell.value.lower():
                datas.append((cell.coordinate, cell.value))
    return datas

def detectar_duplicatas(sheet):
    valores = {}
    duplicatas = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                if cell.value in valores:
                    duplicatas.append((cell.coordinate, cell.value))
                else:
                    valores[cell.value] = cell.coordinate
    return duplicatas

def criar_tabela_em_planilha(sheet, num_linhas, num_colunas, cabecalhos):
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        sheet.cell(row=1, column=col_idx, value=cabecalho)
    
    ref = f"A1:{chr(64 + num_colunas)}{num_linhas}"
    tabela = Table(displayName="Tabela1", ref=ref)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    sheet.add_table(tabela)

def aplicar_formatacao_padronizada(sheet, nome_fonte, tamanho_fonte, negrito, italico, sublinhado, alinhamento_horizontal, alinhamento_vertical, espaco_linhas):
    fonte = Font(name=nome_fonte, size=tamanho_fonte, bold=negrito, italic=italico, underline='single' if sublinhado else None)
    alinhamento = Alignment(horizontal=alinhamento_horizontal, vertical=alinhamento_vertical)
    borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = fonte
            cell.alignment = alinhamento
            cell.border = borda_fina
    
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = espaco_linhas

def processar_arquivos(caminhos_arquivos, nome_planilha, novo_titulo, aplicar_tabela, tabela_linhas, tabela_colunas, cabecalhos, nome_fonte, tamanho_fonte, negrito, italico, sublinhado, alinhamento_horizontal, alinhamento_vertical, espaco_linhas):
    for caminho_arquivo in caminhos_arquivos:
        try:
            wb = load_workbook(caminho_arquivo)
            if nome_planilha not in wb.sheetnames:
                print(f"Planilha '{nome_planilha}' não encontrada em '{caminho_arquivo}'.")
                continue
            
            sheet = wb[nome_planilha]

            # Identificar e editar títulos
            titulos = identificar_e_editar_titulos(sheet, novo_titulo)
            if titulos:
                print(f"Títulos identificados e editados em '{caminho_arquivo}': {titulos}")

            # Verificar datas
            datas = verificar_datas(sheet)
            if datas:
                print(f"Datas encontradas em '{caminho_arquivo}': {datas}")

            # Detectar duplicatas
            duplicatas = detectar_duplicatas(sheet)
            if duplicatas:
                print(f"Duplicatas encontradas em '{caminho_arquivo}': {duplicatas}")

            if aplicar_tabela:
                criar_tabela_em_planilha(sheet, tabela_linhas, tabela_colunas, cabecalhos)

            aplicar_formatacao_padronizada(sheet, nome_fonte, tamanho_fonte, negrito, italico, sublinhado, alinhamento_horizontal, alinhamento_vertical, espaco_linhas)
            
            wb.save(caminho_arquivo)
            print(f"Formatação e tabela aplicadas ao arquivo '{caminho_arquivo}'.")

        except Exception as e:
            print(f"Erro ao processar o arquivo {caminho_arquivo}: {e}")

def main():
    caminhos_arquivos = input("Digite os caminhos dos arquivos Excel, separados por vírgula: ").strip().split(',')
    nome_planilha = input("Digite o nome da planilha que deseja modificar: ").strip()
    
    novo_titulo = input("Digite o novo título para substituir os títulos existentes: ").strip()
    
    aplicar_tabela = input("Deseja adicionar uma tabela na planilha? (s/n): ").strip().lower() == 's'
    
    tabela_linhas = 0
    tabela_colunas = 0
    cabecalhos = []
    
    if aplicar_tabela:
        tabela_linhas = int(input("Digite o número de linhas para a tabela: "))
        tabela_colunas = int(input("Digite o número de colunas para a tabela: "))
        cabecalhos = input(f"Digite os cabeçalhos das colunas, separados por vírgula: ").strip().split(',')

    nome_fonte = input("Digite o nome da fonte que deseja aplicar: ").strip()
    tamanho_fonte = int(input("Digite o tamanho da fonte: "))
    negrito = input("Deseja aplicar negrito? (s/n): ").strip().lower() == 's'
    italico = input("Deseja aplicar itálico? (s/n): ").strip().lower() == 's'
    sublinhado = input("Deseja aplicar sublinhado? (s/n): ").strip().lower() == 's'
    alinhamento_horizontal = input("Digite o alinhamento horizontal (center, left, right): ").strip().lower()
    alinhamento_vertical = input("Digite o alinhamento vertical (center, top, bottom): ").strip().lower()
    
    # Adicionando validação para a altura das linhas
    while True:
        try:
            espaco_linhas_input = input("Digite a altura das linhas para espaçamento (deixe em branco para manter o padrão): ").strip()
            espaco_linhas = float(espaco_linhas_input) if espaco_linhas_input else 15.0  # Valor padrão
            break
        except ValueError:
            print("Valor inválido para altura das linhas. Tente novamente.")

    processar_arquivos(caminhos_arquivos, nome_planilha, novo_titulo, aplicar_tabela, tabela_linhas, tabela_colunas, cabecalhos, nome_fonte, tamanho_fonte, negrito, italico, sublinhado, alinhamento_horizontal, alinhamento_vertical, espaco_linhas)

if __name__ == "__main__":
    main()
