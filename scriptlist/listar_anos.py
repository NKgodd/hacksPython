import os
from openpyxl import load_workbook
import re

def listar_anos_em_excel(pasta, anos_procurados):
    # Verifica se o ano procurado está no intervalo esperado
    anos_procurados = {str(ano) for ano in anos_procurados}

    for root, dirs, files in os.walk(pasta):
        for file in files:
            if file.endswith('.xlsx'):
                caminho_arquivo = os.path.join(root, file)
                try:
                    wb = load_workbook(caminho_arquivo, data_only=True)
                    ano_posicoes = {ano: [] for ano in anos_procurados}

                    for sheet in wb.worksheets:
                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            for col_idx, cell in enumerate(row, start=1):
                                if isinstance(cell, str):
                                    encontrados = re.findall(r'\b(200\d|201\d|2020|2021)\b', cell)
                                    for ano in encontrados:
                                        if ano in ano_posicoes:
                                            ano_posicoes[ano].append(f'Sheet: {sheet.title}, Row: {row_idx}, Column: {col_idx}')
                                elif isinstance(cell, int):
                                    ano = str(cell)
                                    if ano in ano_posicoes:
                                        ano_posicoes[ano].append(f'Sheet: {sheet.title}, Row: {row_idx}, Column: {col_idx}')
                    
                    if any(ano_posicoes.values()):
                        print(f'\nArquivo: {caminho_arquivo}')
                        for ano, posicoes in ano_posicoes.items():
                            if posicoes:
                                print(f'Ano {ano} encontrado em:')
                                for posicao in posicoes:
                                    print(f'  - {posicao}')
                except Exception as e:
                    print(f'Erro ao abrir {caminho_arquivo}: {e}')

# Lista dos anos que você deseja buscar
anos_procurados = list(range(2001, 2022))

# Passe o caminho da pasta onde os arquivos Excel estão localizados
listar_anos_em_excel(r'C:\Users\Adm Manhã\Documents', anos_procurados)
