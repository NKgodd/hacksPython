import os
from openpyxl import load_workbook
import re

def verificar_datas_erradas(pasta):
    for root, dirs, files in os.walk(pasta):
        for file in files:
            if file.endswith('.xlsx'):
                ano_arquivo = re.search(r'(\d{4})', file)
                if ano_arquivo:
                    ano_arquivo = ano_arquivo.group(1)
                    caminho_arquivo = os.path.join(root, file)
                    try:
                        wb = load_workbook(caminho_arquivo, data_only=True)
                        for sheet in wb.worksheets:
                            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                                for col_idx, cell in enumerate(row, start=1):
                                    if isinstance(cell, str):
                                        # Verifica se a célula contém uma data no formato YYYY-MM-DD ou outras variantes
                                        datas_encontradas = re.findall(r'\b(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})\b', cell)
                                        for data in datas_encontradas:
                                            ano_data = re.search(r'\d{4}', data).group()
                                            if ano_data != ano_arquivo:
                                                print(f'Arquivo: {caminho_arquivo}')
                                                print(f'Data errada encontrada: {data} na Planilha: {sheet.title}, Linha: {row_idx}, Coluna: {col_idx}')
                                    elif isinstance(cell, (int, float)):
                                        # Verifica se a célula pode ser uma data representada como número
                                        if 1900 <= cell <= 2099:  # Ajuste conforme necessário
                                            ano_data = str(int(cell))
                                            if ano_data != ano_arquivo:
                                                print(f'Arquivo: {caminho_arquivo}')
                                                print(f'Data errada encontrada: {ano_data} na Planilha: {sheet.title}, Linha: {row_idx}, Coluna: {col_idx}')
                    except Exception as e:
                        print(f'Erro ao abrir {caminho_arquivo}: {e}')

def main():
    pasta = input('Digite o caminho da pasta onde estão os arquivos Excel: ')
    verificar_datas_erradas(pasta)

if __name__ == '__main__':
    main()
