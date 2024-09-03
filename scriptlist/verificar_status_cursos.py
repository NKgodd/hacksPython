import os
from openpyxl import load_workbook

def localizar_status_curso(pasta, status_desistente, status_concluinte, status_aprovado):
    for root, dirs, files in os.walk(pasta):
        for file in files:
            if file.endswith('.xlsx'):
                caminho_arquivo = os.path.join(root, file)
                try:
                    wb = load_workbook(caminho_arquivo, data_only=True)
                    resultados = {
                        'Desistentes': [],
                        'Concluintes': [],
                        'Aprovados': []
                    }

                    for sheet in wb.worksheets:
                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            for col_idx, cell in enumerate(row, start=1):
                                if isinstance(cell, str):
                                    if status_desistente.lower() in cell.lower():
                                        resultados['Desistentes'].append(f'Sheet: {sheet.title}, Row: {row_idx}, Column: {col_idx}')
                                    elif status_concluinte.lower() in cell.lower():
                                        resultados['Concluintes'].append(f'Sheet: {sheet.title}, Row: {row_idx}, Column: {col_idx}')
                                    elif status_aprovado.lower() in cell.lower():
                                        resultados['Aprovados'].append(f'Sheet: {sheet.title}, Row: {row_idx}, Column: {col_idx}')

                    if any(resultados.values()):
                        print(f'\nArquivo: {caminho_arquivo}')
                        for status, posicoes in resultados.items():
                            if posicoes:
                                print(f'{status} encontrado em:')
                                for posicao in posicoes:
                                    print(f'  - {posicao}')
                    
                except Exception as e:
                    print(f'Erro ao abrir {caminho_arquivo}: {e}')

def main():
    caminho_pasta = input("Digite o caminho da pasta onde estão os arquivos Excel: ").strip()
    status_desistente = input("Digite o texto que indica desistentes (por exemplo, 'Desistente'): ").strip()
    status_concluinte = input("Digite o texto que indica concluintes (por exemplo, 'Concluinte'): ").strip()
    status_aprovado = input("Digite o texto que indica aprovados (por exemplo, 'Aprovado'): ").strip()
    
    localizar_status_curso(caminho_pasta, status_desistente, status_concluinte, status_aprovado)

if __name__ == "__main__":
    main()
